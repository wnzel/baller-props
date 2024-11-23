import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font

from datetime import datetime  # Import datetime module

import requests
import os

load_dotenv()
BACKEND_URL = os.getenv("BACKEND_URL")

data = None
res = requests.get(BACKEND_URL)

if res.status_code == 200:
    data = res.json()
else:
    print(f"Failed to get data: {res.status_code}")
    data = None

# Prepare a list to hold all the rows for the spreadsheet
spreadsheet_data = []

# Loop through each game and player
for game_title in data:
    for team in data[game_title]:
        for player in data[game_title][team]:
            lines = data[game_title][team][player]["lines"]
            for market in lines:
                # Get the projected line for the market
                projected_line = lines[market]
                
                # Initialize a counter for times hit over
                times_hit_over = 0
                game_stats = []
                # Loop through the player's last 10 game stats
                for game in data[game_title][team][player]["last_10_game_stats"]:
                    game_stats.append(game[market])
                    if game[market] > projected_line:
                        times_hit_over += 1
                
                # Calculate the hit percentage
                hit_percentage = (times_hit_over / 10) * 100
                expected_value = sum(game_stats) / len(game_stats) if game_stats else 0

                # Add the data for this entry to the list
                spreadsheet_data.append({
                    "Game": game_title,
                    "Player Name": player,
                    "Market": market,
                    "Line": projected_line,
                    "Projection": expected_value,
                    "Times Hit Over": times_hit_over,
                    "Times Hit Under": 10 - times_hit_over,
                    "Hit Percentage": round(hit_percentage, 2)
                })

# Create a pandas DataFrame from the collected data
df = pd.DataFrame(spreadsheet_data)

# 1. **Sort the DataFrame by 'hit percentage last 10' in descending order (highest to lowest)**
df_sorted_desc = df.sort_values(by="Hit Percentage", ascending=False)

current_date = datetime.now().strftime("%m.%d.%Y")  # Format as MM-DD-YYYY

# Save the sorted DataFrame (highest to lowest hit percentage) to an Excel file
excel_file = f"Over_{current_date}.xlsx"
df_sorted_desc.to_excel(excel_file, index=False)  # Save as Excel

# 2. **Sort the DataFrame by 'hit percentage last 10' in ascending order (lowest to highest)**
df_sorted_asc = df.sort_values(by="Hit Percentage", ascending=True)

# Save the sorted DataFrame (lowest to highest hit percentage) to a separate Excel file
excel_file_asc = f"Under_{current_date}.xlsx"
df_sorted_asc.to_excel(excel_file_asc, index=False)  # Save as Excel

# Apply the same formatting to both files using openpyxl

# Function to apply formatting
def apply_formatting(file_name, df):
    # Load the workbook using openpyxl to apply custom formatting
    wb = load_workbook(file_name)
    ws = wb.active

    # Set column widths for better readability
    column_widths = {
        "A": 20,  
        "B": 20,  
        "C": 15,  
        "D": 20,  
        "E": 20,  
        "F": 25,  
        "G": 25,  
        "H": 25   
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define the range for the "hit percentage last 10" column (F column)
    hit_percentage_range = "H2:H" + str(len(df) + 1)  # Adjust for the number of rows

    # Create a ColorScaleRule to apply a color scale (gradient)
    color_scale = ColorScaleRule(
        start_type='min', start_color='FF9C9C',   # White for minimum (0%)
        mid_type='percentile', mid_value=50, mid_color='FFFFFF',  # Yellow for midpoint (50%)
        end_type='max', end_color='1DCB3A'        # Green for maximum (100%)
    )

    # Apply the color scale to the range
    ws.conditional_formatting.add(hit_percentage_range, color_scale)

    # Center-align all cells in the worksheet
    for row in ws.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add a light grey border to each cell
            thin_border = Border(
                left=Side(border_style="thin", color="D3D3D3"),  # Light grey color
                right=Side(border_style="thin", color="D3D3D3"),
                top=Side(border_style="thin", color="D3D3D3"),
                bottom=Side(border_style="thin", color="D3D3D3")
            )
            cell.border = thin_border
        
    # Apply the blue background with yellow text to the header row (first row)
    header_fill = PatternFill(start_color="DADADA", end_color="DADADA", fill_type="solid")  # Blue
    header_font = Font(color="000000", bold=True)  # Yellow text

    for cell in ws[1]:  # Apply to the first row (header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Save the final Excel file with color formatting and centered text
    wb.save(file_name)

# Apply the formatting to the original sorted file (highest to lowest)
apply_formatting(excel_file, df_sorted_desc)

# Apply the formatting to the second sorted file (lowest to highest)
apply_formatting(excel_file_asc, df_sorted_asc)

print(f"Spreadsheet saved as {excel_file}")
print(f"Spreadsheet saved as {excel_file_asc}")